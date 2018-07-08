"Module for grabbing existing properties and adding new properties"
# coding: utf-8

import openpyxl
from settings import settings

SHEET_NAME = "Properties"
ID = "id"
COLUMNS_TO_KEEP = [
    "id",
    "Location",
    "Price",
    "Property type",
    "Rateable value (RV)",
    "Rooms",
    "href",
    "Floor area",
    "Land area",
    "Date listed",
]


def create_workbook_if_not_present(sheet_dict):
    """Create the trademe workbook if not present with the required sheets.
    
    This creates a new workbook based on the provided dictionary. It will
    create a sheet per dictionary key, with the header row as the list
    of the corresponding key.
    """
    try:
        openpyxl.load_workbook(settings.excel_file_address)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        default_sheet = wb.active

        for sheet_name, columns in sheet_dict.values():
            sheet = wb.create_sheet(sheet_name)
            sheet.append(columns)

        wb.remove_sheet(default_sheet)

        wb.save(settings.excel_file_address)


def get_previous_data_from_excel(sheet, column, header_row=1):
    """Return a column of values from the tm excel spreadsheet as a set."""
    xl_sheet = openpyxl.load_workbook(settings.excel_file_address)[sheet]

    # Ascertain the ID column
    headers = xl_sheet[header_row]
    id_column = [cell.column for cell in headers if cell.value == column][0]

    # Turn IDs into a set
    cells = xl_sheet[id_column]
    output = {cell.value for cell in cells[1:]}
    return output


def save_file(data, sheet, columns):
    xl_wb = openpyxl.load_workbook(settings.excel_file_address)
    xl_sheet = xl_wb[sheet]
    data_list = [[row.get(key, "") for key in columns] for row in data]

    for row in data_list:
        xl_sheet.append(row)

    xl_wb.save(settings.excel_file_address)
